from __future__ import annotations

import json
from itertools import islice
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[2]
NAMES = [
    "小分子-靶点（swiss预测靶点）.xlsx",
    "swiss target prediction预测靶点.xlsx",
    "中药小分子靶点.xlsx",
    "AD_中药_小分子_作用靶点_全量表.xlsx",
]


def main() -> None:
    result = []
    for name in NAMES:
        path = ROOT / name
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        item = {"file": name, "size": path.stat().st_size, "sheets": []}
        for ws in wb.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            sample = []
            if max_row and max_col:
                for row in ws.iter_rows(min_row=1, max_row=min(max_row, 5), values_only=True):
                    sample.append(list(row[: min(max_col, 25)]))
            else:
                ws.reset_dimensions()
                for row in islice(ws.iter_rows(values_only=True), 5):
                    sample.append(list(row[:25]))
                max_row = "dimension metadata unavailable"
                max_col = max((len(row) for row in sample), default=0)
            item["sheets"].append(
                {
                    "name": ws.title,
                    "max_row": max_row,
                    "max_col": max_col,
                    "sample": sample,
                }
            )
        wb.close()
        result.append(item)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
